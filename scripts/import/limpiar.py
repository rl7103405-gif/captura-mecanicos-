#!/usr/bin/env python3
"""
Pipeline de limpieza del historico (REPORTE_DE_MECANICOS.xlsx).

Hace:
  - Estandariza las 300+ variantes de escritura de fallas -> 24 categorias.
  - Corrige nombres de mecanicos inconsistentes (ALEJANDO, MARTN, MARCOPS...),
    manteniendo ALEX y ALEJANDRO como personas DISTINTAS.
  - Limpia horas corruptas (celdas con fecha completa) y descarta duraciones
    imposibles (> 8 h o <= 0).
  - Calcula tiempos promedio por tipo de falla para refinar los estandares.
  - El historico NO tiene hora de reporte: solo se obtiene tiempo de reparacion.

Salidas (en scripts/import/out/):
  - incidencias_historico.json     (listo para importar a Firestore)
  - paros.json                     (Reporte de Paros MQ)
  - tiempos_estandar_refinados.json
  - reporte_limpieza.md            (resumen legible)

Uso:  python3 scripts/import/limpiar.py <ruta_al_xlsx>
"""
import sys, os, json, re, datetime
from collections import Counter, defaultdict

import openpyxl

AQUI = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(AQUI, 'out')
ANIO = 2026  # el historico va de marzo a junio (paros confirma 2026)

# ---- Catalogo (24 categorias + min estandar) ----
with open(os.path.join(AQUI, '..', 'catalogo_fallas.json'), encoding='utf-8') as f:
    CATALOGO = json.load(f)
POR_CODIGO = {c['codigo']: c for c in CATALOGO}

# ---- Normalizacion de nombres de mecanicos ----
# Alex y Alejandro son DISTINTOS. Solo corregimos errores de dedo.
MAP_MECANICOS = {
    'ALEJANDO': 'ALEJANDRO',
    'MARTN': 'MARTIN',
    'MARCOPS': 'MARCOS',
    'MARCO': 'MARCOS',
}

def norm_mecanico(s):
    if not s:
        return None
    n = ' '.join(str(s).strip().upper().split())
    return MAP_MECANICOS.get(n, n)

# ---- Clasificador de fallas: reglas ordenadas (gana la primera) ----
# Cada regla es (palabras_clave, codigo). Orden de especifico -> general.
REGLAS = [
    (r'NO DA TALLA', '20'),
    (r'CILINDRO', '12'),
    (r'PUNZON|PUZON', '02'),
    (r'PLATINA|PLATINER', '17'),
    (r'SELECTOR|SLECTOR|SLYDER|SLIDER|SLECTORE', '14'),
    (r'TECLA', '14'),
    (r'AGUJA|GUJA|DESPAT', '01'),
    (r'BANDA', '15'),
    (r'CUCHILL|CUHCILL|CICHILL|CUHILL|NO CORTA|NO CORA|NO CARTA|NO CORTALA|CORTE|CORTO FLECHA|FILO|SIERRA|CIERRA', '03'),
    (r'VANIZ|VANISA', '07'),
    (r'BORDAD|BORADD', '08'),
    (r'VALVULA', '05'),
    (r'ALIMENTAD', '11'),
    (r'TRANSFER|TRANFER', '21'),
    (r'CONTAMINA', '22'),
    (r'HEBRA', '13'),
    (r'ELASTICO|LICRA|LYCRA|\bLIGA\b|ELEATICO|LIFA', '06'),
    (r'FELPA|LISO', '04'),
    (r'ERROR|DISPARO|DIS[AÁ]RO|CODIGO|CODIDO|DISMINUCION', '09'),
    (r'ACTUADOR|PICKER|PICKET|BALANCIN|ALDABA|PISTON|TRIANGULO', '23'),
    (r'AIRE|FUGA|SUCCION', '19'),
    (r'ELECTRIC|FUENTE|MOTOR|SE APAGO', '18'),
    (r'TALLA|CADENA|GIROS|VALETA|AMARRE', '16'),
    (r'PICA|TROZA|BOTA|REVIENTA|SUELTA|PUNTO ABIERTO|\bRAYA\b|NO SUELTA|TEJIDO|ABRE|SE CARG|CARGA', '10'),
]
REGLAS = [(re.compile(p), c) for p, c in REGLAS]

def clasificar(texto):
    if texto is None:
        return '24'
    t = ' '.join(str(texto).strip().upper().split())
    if not t or t.isdigit():
        return '24'
    for rx, cod in REGLAS:
        if rx.search(t):
            return cod
    return '24'

# ---- Utilidades de tiempo ----
MESES = {'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
         'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12}

def fecha_de_hoja(titulo):
    t = titulo.strip().upper()
    m = re.match(r'\s*(\d{1,2})\s*([A-ZÑ]+)', t)
    if not m:
        return None
    dia = int(m.group(1))
    mes = MESES.get(m.group(2))
    if not mes:
        return None
    try:
        return datetime.date(ANIO, mes, dia)
    except ValueError:
        return None

def a_hora(celda):
    """Devuelve datetime.time o None. Corrige celdas con fecha completa."""
    if celda is None:
        return None
    if isinstance(celda, datetime.datetime):  # fecha completa corrupta -> usar la hora
        return celda.time()
    if isinstance(celda, datetime.time):
        return celda
    return None

def combinar(fecha, hora):
    if fecha is None or hora is None:
        return None
    return datetime.datetime.combine(fecha, hora)

# ---- Procesar hojas diarias (incidencias de mecanicos) ----
def procesar(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    incidencias = []
    descartes = Counter()
    sin_clasificar = Counter()
    dur_por_cat = defaultdict(list)
    contador = 0

    for ws in wb.worksheets:
        if 'PARO' in ws.title.upper():
            continue
        fecha = fecha_de_hoja(ws.title)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row or all(c is None for c in row):
                continue
            turno = row[0] if len(row) > 0 else None
            mq = row[1] if len(row) > 1 else None
            falla_txt = row[2] if len(row) > 2 else None
            mec = row[3] if len(row) > 3 else None
            h_ini = a_hora(row[4] if len(row) > 4 else None)
            h_fin = a_hora(row[5] if len(row) > 5 else None)

            if mq is None or not isinstance(mq, (int, float)):
                descartes['sin maquina'] += 1
                continue
            if h_ini is None or h_fin is None:
                descartes['hora faltante/ilegible'] += 1
                continue

            ini = combinar(fecha, h_ini)
            fin = combinar(fecha, h_fin)
            if ini is None or fin is None:
                descartes['fecha de hoja ilegible'] += 1
                continue
            if fin <= ini:  # cruza medianoche -> +1 dia
                fin += datetime.timedelta(days=1)
            dur = round((fin - ini).total_seconds() / 60)

            if dur <= 0:
                descartes['duracion cero'] += 1
                continue
            if dur > 480:  # > 8 h imposible para una reparacion
                descartes['duracion > 8 h'] += 1
                continue

            cod = clasificar(falla_txt)
            if cod == '24' and falla_txt:
                sin_clasificar[' '.join(str(falla_txt).upper().split())] += 1
            dur_por_cat[cod].append(dur)

            contador += 1
            cat = POR_CODIGO[cod]
            incidencias.append({
                'folio': f'HIST-{contador:05d}',
                'maquina': int(mq),
                'estado': 'cerrada',
                'fuente': 'historico',
                'turno': int(turno) if isinstance(turno, (int, float)) else None,
                'tejedorNombre': None,
                'mecanicoNombre': norm_mecanico(mec),
                'descripcion': str(falla_txt).strip() if falla_txt else '',
                'tipoFalla': {'codigo': cat['codigo'], 'falla': cat['falla'],
                              'minEstandar': cat['minEstandar']},
                'reportadaEn': None,                  # el historico NO tiene hora de reporte
                'aceptadaEn': ini.isoformat(),
                'cerradaEn': fin.isoformat(),
                'tiempoRespuestaMin': None,            # no se puede calcular sin reporte
                'tiempoReparacionMin': dur,
                'tiempoMuertoMin': dur,                # aprox: solo abarca la reparacion
            })

    paros = procesar_paros(wb)
    return incidencias, paros, descartes, sin_clasificar, dur_por_cat

# ---- Procesar hoja de paros ----
def procesar_paros(wb):
    paros = []
    ws = None
    for w in wb.worksheets:
        if 'PARO' in w.title.upper():
            ws = w
            break
    if ws is None:
        return paros
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or all(c is None for c in row):
            continue
        reporto = row[0]
        f_paro = row[1]
        mq = row[2]
        folio = row[3]
        motivo = row[4]
        operador = row[5]
        h_paro = a_hora(row[6])
        f_arr = row[7]
        h_arr = a_hora(row[8])
        if mq is None or not isinstance(mq, (int, float)):
            continue
        fp = f_paro.date() if isinstance(f_paro, datetime.datetime) else None
        fa = f_arr.date() if isinstance(f_arr, datetime.datetime) else fp
        paroEn = combinar(fp, h_paro)
        arrEn = combinar(fa, h_arr)
        dur = None
        if paroEn and arrEn:
            if arrEn < paroEn:
                arrEn += datetime.timedelta(days=1)
            dur = round((arrEn - paroEn).total_seconds() / 60)
            if dur <= 0 or dur > 24 * 60:  # paros: hasta 24 h
                dur = None
        n += 1
        paros.append({
            'folio': int(folio) if isinstance(folio, (int, float)) else (str(folio) if folio else f'PARO-{n}'),
            'maquina': int(mq),
            'fuente': 'paros',
            'motivo': str(motivo).strip() if motivo else '',
            'operador': str(operador).strip() if operador else None,
            'reporto': str(reporto).strip() if reporto else None,
            'paroEn': paroEn.isoformat() if paroEn else None,
            'arranqueEn': arrEn.isoformat() if arrEn else None,
            'duracionMin': dur,
        })
    return paros

def main():
    if len(sys.argv) < 2:
        print('Uso: python3 limpiar.py <ruta_al_xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    os.makedirs(OUT, exist_ok=True)

    inc, paros, descartes, sin_clas, dur_por_cat = procesar(path)

    # Tiempos estandar refinados (promedio real por categoria).
    refinados = []
    for c in CATALOGO:
        durs = dur_por_cat.get(c['codigo'], [])
        prom = round(sum(durs) / len(durs)) if durs else None
        refinados.append({
            'codigo': c['codigo'], 'falla': c['falla'],
            'minEstandarActual': c['minEstandar'],
            'minPromedioReal': prom, 'n': len(durs),
        })

    with open(os.path.join(OUT, 'incidencias_historico.json'), 'w', encoding='utf-8') as f:
        json.dump(inc, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, 'paros.json'), 'w', encoding='utf-8') as f:
        json.dump(paros, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, 'tiempos_estandar_refinados.json'), 'w', encoding='utf-8') as f:
        json.dump(refinados, f, ensure_ascii=False, indent=2)

    # Reporte legible.
    por_mec = Counter(i['mecanicoNombre'] for i in inc)
    por_cat = Counter(i['tipoFalla']['falla'] for i in inc)
    lineas = []
    lineas.append('# Reporte de limpieza del historico\n')
    lineas.append(f'- Incidencias validas importables: **{len(inc)}**')
    lineas.append(f'- Paros importables: **{len(paros)}**')
    lineas.append(f'- Descartes: **{sum(descartes.values())}**\n')
    lineas.append('## Descartes por motivo')
    for k, v in descartes.most_common():
        lineas.append(f'- {k}: {v}')
    lineas.append('\n## Incidencias por mecanico (nombres normalizados)')
    for k, v in por_mec.most_common():
        lineas.append(f'- {k}: {v}')
    lineas.append('\n## Incidencias por categoria de falla')
    for k, v in por_cat.most_common():
        lineas.append(f'- {k}: {v}')
    lineas.append('\n## Tiempos estandar: actual vs promedio real')
    lineas.append('| Cod | Falla | Estandar | Prom. real | n |')
    lineas.append('|---|---|---|---|---|')
    for r in refinados:
        lineas.append(f"| {r['codigo']} | {r['falla']} | {r['minEstandarActual']} | "
                      f"{r['minPromedioReal'] if r['minPromedioReal'] is not None else '—'} | {r['n']} |")
    if sin_clas:
        lineas.append(f'\n## Textos no clasificados -> "Otro" (revisar, {sum(sin_clas.values())} casos)')
        for k, v in sin_clas.most_common(30):
            lineas.append(f'- ({v}) {k}')
    with open(os.path.join(OUT, 'reporte_limpieza.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lineas) + '\n')

    print(f'OK -> {len(inc)} incidencias, {len(paros)} paros. Salidas en {OUT}')
    print(f'Descartes: {sum(descartes.values())} | a "Otro": {sum(sin_clas.values())}')

if __name__ == '__main__':
    main()
